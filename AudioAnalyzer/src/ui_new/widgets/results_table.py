"""
Виджет таблицы результатов с контекстным меню.

Назначение:
- Отображение результатов обработки
- Контекстное меню для действий
- Копирование значений
- Открытие файлов

Использование:
--------------
>>> from ui_new.widgets.results_table import ResultsTable
>>> 
>>> table = ResultsTable()
>>> table.add_result(result_row)
"""
from __future__ import annotations

import logging
import os
import subprocess
import sys
from typing import Any, Dict, List, Optional

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QAction, QKeySequence
from PySide6.QtWidgets import (
    QApplication,
    QMenu,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
)

logger = logging.getLogger("ui_new.widgets.results_table")


class ResultsTable(QTableWidget):
    """Таблица результатов с контекстным меню.
    
    Предоставляет:
    - Отображение результатов обработки
    - Контекстное меню для действий
    - Копирование значений в буфер обмена
    - Открытие файлов и папок
    
    Сигналы:
    ----------
    file_opened : Signal(str)
        Файл открыт
    file_analyzed : Signal(str)
        Запрошен анализ файла
    """
    
    # Сигналы
    file_opened = Signal(str)
    file_analyzed = Signal(str)
    
    def __init__(self, parent=None):
        """Инициализация таблицы."""
        super().__init__(parent)
        
        # Настройка таблицы
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.setSortingEnabled(True)
        self.setAlternatingRowColors(True)
        
        # Включаем контекстное меню
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        
        # Кэш результатов для быстрого доступа
        self._results_cache: List[Any] = []
        
        logger.debug("ResultsTable initialized")
    
    def _show_context_menu(self, pos) -> None:
        """Показать контекстное меню.
        
        Параметры:
        ----------
        pos : QPoint
            Позиция клика
        """
        # Получаем выбранные строки
        selected_rows = self._get_selected_rows()
        
        if not selected_rows:
            return
        
        menu = QMenu(self)
        
        # Действия для одной строки
        if len(selected_rows) == 1:
            row = selected_rows[0]
            path = self._get_path_for_row(row)
            
            # Открыть файл
            open_action = QAction("▶ Открыть файл", self)
            open_action.triggered.connect(lambda: self._open_file(path))
            menu.addAction(open_action)
            
            # Открыть папку
            folder_action = QAction("📂 Открыть папку", self)
            folder_action.triggered.connect(lambda: self._open_folder(path))
            menu.addAction(folder_action)
            
            menu.addSeparator()
            
            # Анализировать спектр
            analyze_action = QAction("📊 Анализ спектра", self)
            analyze_action.triggered.connect(lambda: self._analyze_file(path))
            menu.addAction(analyze_action)
            
            menu.addSeparator()
        
        # Действия для множества строк
        copy_values_action = QAction("📋 Копировать значения", self)
        copy_values_action.triggered.connect(self._copy_selected_values)
        menu.addAction(copy_values_action)
        
        copy_row_action = QAction("📋 Копировать строку", self)
        copy_row_action.triggered.connect(self._copy_selected_rows)
        menu.addAction(copy_row_action)
        
        menu.addSeparator()
        
        # Экспорт выбранных
        export_action = QAction("📄 Экспорт выбранных...", self)
        export_action.triggered.connect(self._export_selected)
        menu.addAction(export_action)
        
        # Показываем меню
        menu.exec(self.viewport().mapToGlobal(pos))
    
    def _get_selected_rows(self) -> List[int]:
        """Получить выбранные строки.
        
        Возвращает:
        -----------
        List[int]
            Список номеров строк
        """
        rows = set()
        for item in self.selectedItems():
            rows.add(item.row())
        return sorted(rows)
    
    def _get_path_for_row(self, row: int) -> str:
        """Получить путь для строки.
        
        Параметры:
        ----------
        row : int
            Номер строки
            
        Возвращает:
        -----------
        str
            Путь к файлу
        """
        # Путь обычно в последней колонке
        path_item = self.item(row, self.columnCount() - 1)
        if path_item:
            return path_item.text()
        return ""
    
    def _open_file(self, path: str) -> None:
        """Открыть файл в системном приложении.
        
        Параметры:
        ----------
        path : str
            Путь к файлу
        """
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "Ошибка", f"Файл не найден:\n{path}")
            return
        
        try:
            if sys.platform == 'win32':
                os.startfile(path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', path])
            else:
                subprocess.run(['xdg-open', path])
            
            self.file_opened.emit(path)
            logger.info(f"Opened file: {path}")
            
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось открыть файл:\n{e}")
            logger.error(f"Error opening file: {e}")
    
    def _open_folder(self, path: str) -> None:
        """Открыть папку в файловом менеджере.
        
        Параметры:
        ----------
        path : str
            Путь к файлу
        """
        if not path:
            return
        
        folder = os.path.dirname(path) if os.path.isfile(path) else path
        
        if not folder or not os.path.exists(folder):
            QMessageBox.warning(self, "Ошибка", f"Папка не найдена:\n{folder}")
            return
        
        try:
            if sys.platform == 'win32':
                os.startfile(folder)
            elif sys.platform == 'darwin':
                subprocess.run(['open', folder])
            else:
                subprocess.run(['xdg-open', folder])
            
            logger.info(f"Opened folder: {folder}")
            
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось открыть папку:\n{e}")
            logger.error(f"Error opening folder: {e}")
    
    def _analyze_file(self, path: str) -> None:
        """Запросить анализ файла.
        
        Параметры:
        ----------
        path : str
            Путь к файлу
        """
        if path:
            self.file_analyzed.emit(path)
    
    def _copy_selected_values(self) -> None:
        """Копировать выбранные значения в буфер обмена."""
        selected = self.selectedItems()
        if not selected:
            return
        
        values = []
        for item in selected:
            values.append(item.text())
        
        text = "\t".join(values)
        QApplication.clipboard().setText(text)
    
    def _copy_selected_rows(self) -> None:
        """Копировать выбранные строки в буфер обмена."""
        rows = self._get_selected_rows()
        if not rows:
            return
        
        lines = []
        for row in rows:
            values = []
            for col in range(self.columnCount()):
                item = self.item(row, col)
                values.append(item.text() if item else "")
            lines.append("\t".join(values))
        
        text = "\n".join(lines)
        QApplication.clipboard().setText(text)
        
        # Показываем сообщение
        logger.info(f"Copied {len(rows)} rows to clipboard")
    
    def _export_selected(self) -> None:
        """Экспортировать выбранные строки."""
        rows = self._get_selected_rows()
        if not rows:
            QMessageBox.information(self, "Экспорт", "Нет выбранных строк")
            return
        
        # Формируем данные
        data = []
        headers = []
        
        # Заголовки
        for col in range(self.columnCount()):
            header = self.horizontalHeaderItem(col)
            headers.append(header.text() if header else f"Col{col}")
        
        # Данные
        for row in rows:
            row_data = {}
            for col in range(self.columnCount()):
                item = self.item(row, col)
                row_data[headers[col]] = item.text() if item else ""
            data.append(row_data)
        
        # Диалог сохранения
        from PySide6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт результатов",
            "results_export.xlsx",
            "Excel (*.xlsx);;CSV (*.csv);;JSON (*.json)"
        )
        
        if not file_path:
            return
        
        # Экспорт
        try:
            if file_path.endswith('.json'):
                import json
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            
            elif file_path.endswith('.csv'):
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(data)
            
            elif file_path.endswith('.xlsx'):
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    
                    # Заголовки
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Данные
                    for row_idx, row_data in enumerate(data, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))
                    
                    wb.save(file_path)
                    
                except ImportError:
                    QMessageBox.warning(
                        self,
                        "Ошибка",
                        "Для экспорта в Excel необходимо установить openpyxl"
                    )
                    return
            
            QMessageBox.information(
                self,
                "Экспорт",
                f"Экспортировано {len(data)} строк в:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка экспорта:\n{e}")
            logger.error(f"Export error: {e}")


# =============================================================================
# ЭКСПОРТ
# =============================================================================

__all__ = ["ResultsTable"]
