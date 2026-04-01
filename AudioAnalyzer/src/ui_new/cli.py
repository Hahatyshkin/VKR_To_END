#!/usr/bin/env python3
"""
CLI интерфейс для AudioAnalyzer.

Назначение:
- Batch-обработка аудиофайлов из командной строки
- Анализ аудиофайлов
- Генерация отчётов

Использование:
--------------
# Обработать папку с WAV файлами
python -m ui_new.cli process -i ./input -o ./output

# Обработать один файл
python -m ui_new.cli process -i audio.wav -o ./output

# Показать информацию о файле
python -m ui_new.cli analyze -f audio.wav

# Использовать только определённые методы
python -m ui_new.cli process -i ./input -o ./output -m fwht fft dct
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Any

# Добавляем путь к src если нужно
src_path = Path(__file__).parent.parent
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("ui_new.cli")


# =============================================================================
# КОМАНДА PROCESS
# =============================================================================

def cmd_process(args: argparse.Namespace) -> int:
    """Обработать аудиофайлы.
    
    Параметры:
    ----------
    args : argparse.Namespace
        Аргументы командной строки
        
    Возвращает:
    -----------
    int
        Код возврата (0 = успех)
    """
    input_path = args.input
    output_dir = args.output
    methods = args.methods
    config_file = args.config
    parallel = args.parallel
    workers = args.workers
    
    # Загружаем конфигурацию
    settings = _load_config(config_file) if config_file else {}
    
    # Переопределяем настройки из аргументов
    if args.bitrate:
        settings['bitrate'] = args.bitrate
    if args.block_size:
        settings['block_size'] = args.block_size
    
    # Определяем список файлов
    files = _collect_files(input_path)
    
    if not files:
        print(f"❌ Файлы не найдены: {input_path}")
        return 1
    
    print(f"📁 Найдено файлов: {len(files)}")
    print(f"📂 Выходная директория: {output_dir}")
    print(f"⚙️  Методы: {', '.join(methods) if methods else 'все'}")
    print(f"🔄 Параллельная обработка: {'да' if parallel else 'нет'}")
    print()
    
    # Создаём выходную директорию
    os.makedirs(output_dir, exist_ok=True)
    
    # Обрабатываем файлы
    start_time = time.time()
    results = []
    
    for i, file_path in enumerate(files, 1):
        print(f"[{i}/{len(files)}] Обработка: {os.path.basename(file_path)}")
        
        try:
            file_results = _process_file(
                file_path,
                output_dir,
                settings,
                methods,
            )
            results.extend(file_results)
            
            for r in file_results:
                status = "✅" if r.get('success', True) else "❌"
                print(f"  {status} {r.get('variant', 'unknown')}: {r.get('time_sec', 0):.2f}с")
                
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            logger.error(f"Error processing {file_path}: {e}")
    
    elapsed = time.time() - start_time
    
    print()
    print(f"⏱️  Время обработки: {elapsed:.2f}с")
    print(f"📊 Всего результатов: {len(results)}")
    
    # Экспорт результатов
    if args.export:
        export_path = args.export
        _export_results(results, export_path)
        print(f"📄 Результаты экспортированы: {export_path}")
    
    # Выводим JSON если нужно
    if args.json:
        print(json.dumps(results, indent=2, ensure_ascii=False))
    
    return 0


def _collect_files(input_path: str) -> List[str]:
    """Собрать список файлов для обработки.
    
    Параметры:
    ----------
    input_path : str
        Путь к файлу или папке
        
    Возвращает:
    -----------
    List[str]
        Список путей к файлам
    """
    path = Path(input_path)
    
    if path.is_file():
        return [str(path)]
    
    if path.is_dir():
        return [str(f) for f in path.rglob("*.wav")]
    
    return []


def _process_file(
    file_path: str,
    output_dir: str,
    settings: Dict[str, Any],
    methods: Optional[List[str]] = None,
) -> List[Dict]:
    """Обработать один файл.
    
    Параметры:
    ----------
    file_path : str
        Путь к файлу
    output_dir : str
        Выходная директория
    settings : Dict[str, Any]
        Настройки обработки
    methods : Optional[List[str]]
        Список методов для обработки
        
    Возвращает:
    -----------
    List[Dict]
        Результаты обработки
    """
    from processing.audio_ops import (
        fwht_transform_and_mp3,
        fft_transform_and_mp3,
        dct_transform_and_mp3,
        wavelet_transform_and_mp3,
        huffman_like_transform_and_mp3,
        rosenbrock_like_transform_and_mp3,
        standard_convert_to_mp3,
        _compute_metrics_batch,
    )
    
    # Все доступные методы
    all_methods = {
        'standard': lambda: standard_convert_to_mp3(file_path, output_dir, bitrate=settings.get('bitrate', '192k')),
        'fwht': lambda: fwht_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['block_size', 'select_mode', 'keep_energy_ratio', 'sequency_keep_ratio', 'bitrate'])),
        'fft': lambda: fft_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['block_size', 'select_mode', 'keep_energy_ratio', 'sequency_keep_ratio', 'bitrate'])),
        'dct': lambda: dct_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['block_size', 'select_mode', 'keep_energy_ratio', 'sequency_keep_ratio', 'bitrate'])),
        'dwt': lambda: wavelet_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['block_size', 'select_mode', 'keep_energy_ratio', 'sequency_keep_ratio', 'levels', 'bitrate'])),
        'huffman': lambda: huffman_like_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['block_size', 'bitrate', 'mu', 'bits'])),
        'rosenbrock': lambda: rosenbrock_like_transform_and_mp3(file_path, output_dir, **_filter_settings(settings, ['alpha', 'beta', 'bitrate'], {'alpha': 'rosen_alpha', 'beta': 'rosen_beta'})),
    }
    
    # Фильтруем методы
    selected_methods = methods if methods else list(all_methods.keys())
    
    results = []
    
    for method in selected_methods:
        if method not in all_methods:
            print(f"  ⚠️  Неизвестный метод: {method}")
            continue
        
        try:
            start = time.time()
            mp3_path, time_sec = all_methods[method]()
            elapsed = time.time() - start
            
            results.append({
                'variant': method.upper(),
                'path': mp3_path,
                'time_sec': time_sec,
                'size_bytes': os.path.getsize(mp3_path) if os.path.exists(mp3_path) else 0,
                'success': True,
            })
            
        except Exception as e:
            results.append({
                'variant': method.upper(),
                'path': '',
                'time_sec': 0,
                'size_bytes': 0,
                'success': False,
                'error': str(e),
            })
    
    # Вычисляем метрики
    if results:
        items = [(r['variant'], r['path'], r['time_sec']) for r in results if r['success']]
        try:
            metrics = _compute_metrics_batch(file_path, items)
            # Объединяем результаты с метриками
            for r in results:
                for m in metrics:
                    if m.get('variant') == r['variant']:
                        r.update(m)
                        break
        except Exception as e:
            logger.warning(f"Could not compute metrics: {e}")
    
    return results


def _filter_settings(
    settings: Dict[str, Any],
    keys: List[str],
    rename: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """Отфильтровать настройки по ключам.
    
    Параметры:
    ----------
    settings : Dict[str, Any]
        Полные настройки
    keys : List[str]
        Ключи для сохранения
    rename : Optional[Dict[str, str]]
        Карта переименования ключей
        
    Возвращает:
    -----------
    Dict[str, Any]
        Отфильтрованные настройки
    """
    rename = rename or {}
    result = {}
    for key in keys:
        src_key = rename.get(key, key)
        if src_key in settings:
            result[key] = settings[src_key]
    return result


def _load_config(config_path: str) -> Dict[str, Any]:
    """Загрузить конфигурацию из файла.
    
    Параметры:
    ----------
    config_path : str
        Путь к файлу конфигурации
        
    Возвращает:
    -----------
    Dict[str, Any]
        Конфигурация
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.warning(f"Could not load config: {e}")
        return {}


def _export_results(results: List[Dict], export_path: str) -> None:
    """Экспортировать результаты в файл.
    
    Параметры:
    ----------
    results : List[Dict]
        Результаты обработки
    export_path : str
        Путь к файлу экспорта
    """
    ext = os.path.splitext(export_path)[1].lower()
    
    if ext == '.json':
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
    
    elif ext == '.csv':
        import csv
        with open(export_path, 'w', newline='', encoding='utf-8') as f:
            if results:
                writer = csv.DictWriter(f, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)
    
    elif ext == '.xlsx':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            if results:
                # Заголовки
                for col, key in enumerate(results[0].keys(), 1):
                    ws.cell(row=1, column=col, value=key)
                
                # Данные
                for row, r in enumerate(results, 2):
                    for col, key in enumerate(r.keys(), 1):
                        ws.cell(row=row, column=col, value=r[key])
            
            wb.save(export_path)
        except ImportError:
            logger.error("openpyxl not installed, falling back to JSON")
            _export_results(results, export_path.replace('.xlsx', '.json'))
    
    else:
        # По умолчанию JSON
        _export_results(results, export_path + '.json')


# =============================================================================
# КОМАНДА ANALYZE
# =============================================================================

def cmd_analyze(args: argparse.Namespace) -> int:
    """Анализировать аудиофайл.
    
    Параметры:
    ----------
    args : argparse.Namespace
        Аргументы командной строки
        
    Возвращает:
    -----------
    int
        Код возврата
    """
    file_path = args.file
    
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        return 1
    
    try:
        from processing.codecs import load_wav_mono, decode_audio_to_mono, get_audio_meta
        
        # Загружаем файл
        try:
            signal, sr = decode_audio_to_mono(file_path)
        except Exception:
            signal, sr = load_wav_mono(file_path)
        
        # Получаем метаданные
        meta = get_audio_meta(file_path)
        
        # Вычисляем характеристики
        duration = len(signal) / sr
        rms = (signal ** 2).mean() ** 0.5
        peak = abs(signal).max()
        dynamic_range = 20 * (peak / (rms + 1e-10)) if rms > 0 else 0
        
        # Спектральный центроид
        from scipy.fft import rfft, rfftfreq
        fft = rfft(signal)
        freqs = rfftfreq(len(signal), 1/sr)
        magnitudes = abs(fft)
        spectral_centroid = (freqs * magnitudes).sum() / (magnitudes.sum() + 1e-10)
        
        print(f"📁 Файл: {os.path.basename(file_path)}")
        print(f"📂 Путь: {file_path}")
        print(f"📏 Размер: {os.path.getsize(file_path) / (1024*1024):.2f} МБ")
        print()
        print("📊 Аудиохарактеристики:")
        print(f"  Длительность: {duration:.2f} сек")
        print(f"  Частота дискретизации: {sr} Гц")
        print(f"  Каналы: {meta.get('channels', '?')}")
        print(f"  RMS: {rms:.4f}")
        print(f"  Пик: {peak:.4f}")
        print(f"  Динамический диапазон: {dynamic_range:.1f} дБ")
        print(f"  Спектральный центроид: {spectral_centroid:.0f} Гц")
        
        if args.json:
            result = {
                'file': file_path,
                'size_bytes': os.path.getsize(file_path),
                'duration_sec': duration,
                'sample_rate': sr,
                'channels': meta.get('channels'),
                'rms': rms,
                'peak': peak,
                'dynamic_range_db': dynamic_range,
                'spectral_centroid_hz': spectral_centroid,
            }
            print()
            print(json.dumps(result, indent=2, ensure_ascii=False))
        
        return 0
        
    except Exception as e:
        print(f"❌ Ошибка анализа: {e}")
        logger.exception("Analysis error")
        return 1


# =============================================================================
# КОМАНДА RECOMMEND
# =============================================================================

def cmd_recommend(args: argparse.Namespace) -> int:
    """Показать рекомендации по методам.
    
    Параметры:
    ----------
    args : argparse.Namespace
        Аргументы командной строки
        
    Возвращает:
    -----------
    int
        Код возврата
    """
    file_path = args.file
    
    if not os.path.exists(file_path):
        print(f"❌ Файл не найден: {file_path}")
        return 1
    
    try:
        from processing.codecs import decode_audio_to_mono, load_wav_mono
        from scipy.fft import rfft, rfftfreq
        
        # Загружаем файл
        try:
            signal, sr = decode_audio_to_mono(file_path)
        except Exception:
            signal, sr = load_wav_mono(file_path)
        
        # Вычисляем характеристики
        duration = len(signal) / sr
        rms = (signal ** 2).mean() ** 0.5
        peak = abs(signal).max()
        dynamic_range = 20 * (peak / (rms + 1e-10)) if rms > 0 else 0
        
        fft = rfft(signal)
        freqs = rfftfreq(len(signal), 1/sr)
        magnitudes = abs(fft)
        spectral_centroid = (freqs * magnitudes).sum() / (magnitudes.sum() + 1e-10)
        
        # Определяем, похоже ли на речь (эвристика)
        is_speech = spectral_centroid < 4000 and dynamic_range < 30
        
        # Формируем рекомендации
        recommendations = []
        
        # FWHT - быстрый для коротких файлов
        if duration < 10:
            recommendations.append(("fwht", 0.9, "Короткий файл - быстрая обработка"))
        else:
            recommendations.append(("fwht", 0.6, "Базовый метод"))
        
        # FFT - универсальный
        recommendations.append(("fft", 0.8, "Универсальный метод"))
        
        # DCT - хорошее сжатие
        if dynamic_range > 40:
            recommendations.append(("dct", 0.85, "Широкий динамический диапазон"))
        else:
            recommendations.append(("dct", 0.7, "Стандартный метод"))
        
        # DWT - для низкочастотного
        if spectral_centroid < 2000:
            recommendations.append(("dwt", 0.85, "Низкочастотный контент"))
        else:
            recommendations.append(("dwt", 0.65, "Вейвлет-преобразование"))
        
        # Huffman - для речи
        if is_speech:
            recommendations.append(("huffman", 0.9, "Речевой сигнал - оптимизация"))
        else:
            recommendations.append(("huffman", 0.6, "μ-law компандирование"))
        
        # Rosenbrock
        recommendations.append(("rosenbrock", 0.5, "Экспериментальный"))
        
        # Standard
        recommendations.append(("standard", 0.75, "Базовый MP3"))
        
        # Сортируем
        recommendations.sort(key=lambda x: x[1], reverse=True)
        
        print(f"📁 Файл: {os.path.basename(file_path)}")
        print(f"📊 Характеристики:")
        print(f"  Длительность: {duration:.1f} сек")
        print(f"  Динамический диапазон: {dynamic_range:.1f} дБ")
        print(f"  Спектральный центроид: {spectral_centroid:.0f} Гц")
        print(f"  Тип: {'Речь' if is_speech else 'Музыка/Другое'}")
        print()
        print("🎯 Рекомендации методов:")
        
        for method, score, reason in recommendations:
            bar = "█" * int(score * 10) + "░" * (10 - int(score * 10))
            print(f"  {method.upper():12} [{bar}] {score:.0%} - {reason}")
        
        return 0
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        logger.exception("Recommendation error")
        return 1


# =============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# =============================================================================

def main() -> int:
    """Главная функция CLI.
    
    Возвращает:
    -----------
    int
        Код возврата
    """
    parser = argparse.ArgumentParser(
        description="AudioAnalyzer CLI - командная строка для обработки аудио",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Команды')
    
    # Команда process
    process_parser = subparsers.add_parser('process', help='Обработать аудиофайлы')
    process_parser.add_argument('-i', '--input', required=True, help='Входной файл или папка')
    process_parser.add_argument('-o', '--output', required=True, help='Выходная папка')
    process_parser.add_argument('-m', '--methods', nargs='+', help='Методы обработки')
    process_parser.add_argument('-c', '--config', help='Файл конфигурации')
    process_parser.add_argument('--bitrate', help='Битрейт MP3')
    process_parser.add_argument('--block-size', type=int, help='Размер блока')
    process_parser.add_argument('--parallel', action='store_true', help='Параллельная обработка')
    process_parser.add_argument('--workers', type=int, default=4, help='Число потоков')
    process_parser.add_argument('--export', help='Экспорт результатов в файл')
    process_parser.add_argument('--json', action='store_true', help='Вывести результаты в JSON')
    
    # Команда analyze
    analyze_parser = subparsers.add_parser('analyze', help='Анализировать аудиофайл')
    analyze_parser.add_argument('-f', '--file', required=True, help='Аудиофайл')
    analyze_parser.add_argument('--json', action='store_true', help='Вывести в JSON')
    
    # Команда recommend
    recommend_parser = subparsers.add_parser('recommend', help='Рекомендации методов')
    recommend_parser.add_argument('-f', '--file', required=True, help='Аудиофайл')
    
    args = parser.parse_args()
    
    if args.command == 'process':
        return cmd_process(args)
    elif args.command == 'analyze':
        return cmd_analyze(args)
    elif args.command == 'recommend':
        return cmd_recommend(args)
    else:
        parser.print_help()
        return 0


if __name__ == '__main__':
    sys.exit(main())
